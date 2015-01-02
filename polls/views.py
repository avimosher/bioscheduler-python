from django.shortcuts import render, render_to_response
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.template.loader import render_to_string
from django.template import RequestContext
import json
import sys
import io
import os
from Bio.Seq import Seq, reverse_complement
from Bio.SeqRecord import SeqRecord
from Bio.SeqFeature import SeqFeature
from Bio import SeqIO
from Bio import AlignIO
from Bio.Alphabet import IUPAC
import polls.SeqToJSON
import dropbox
from dropbox.client import DropboxOAuth2Flow
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from ast import literal_eval
import traceback
from celery.result import AsyncResult
from mysite.celery import app
from polls.tasks import test2
from subprocess import PIPE,Popen

# Create your views here.

def poll_state(request):
    data={}
    try:
        if 'job' in request.POST:
            job_id=request.POST['job']
        else:
            return HttpResponse('No job id')
        job=AsyncResult(job_id)
        data=job.result or job.state
    except Exception as e:
        print(str(e))
       	traceback.print_exc()
       	raise e
    return HttpResponse(json.dumps(data),content_type='application/json')

def init_work(request):
    try:
        app.set_current()
        job=test2.delay('output')
    except Exception as e:
        print(str(e))
       	traceback.print_exc()
       	raise e
    return HttpResponse(job.id)

def index(request):
    return HttpResponse("Hello, world.  You're at the poll index.")

def detail(request, poll_id):
    return HttpResponse("You're looking at poll %s."%poll_id)

def results(request, poll_id):
    return HttpResponse("You're looking at the results of poll %s."%poll_id)

def vote(request,poll_id):
    return HttpResponse("You're voting on poll %s." % poll_id)

def registerdropbox(request):
	flow=get_dropbox_auth_flow(request.session)
	copyget=request.GET.copy()
	getdict=request.session
	# this converts the item in the URL to the string representation of its byte array.  It either works or errors and isn't necessary
	try:
		copyget.__setitem__('state',literal_eval(copyget['state']))
	except:
		pass
	# this does the same for the item in the session
	getdict['dropbox-auth-csrf-token']=getdict['dropbox-auth-csrf-token'].decode(encoding="UTF-8")
	try:
		access_token, user_id, url_state=flow.finish(copyget)
		request.session['access_token']=access_token
		request.session['user_id']=user_id
	except DropboxOAuth2Flow.BadRequestException as e:
		http_status(400)
	except DropboxOAuth2Flow.BadStateException as e:
		# Start the auth flow again.
		return HttpResponseRedirect("http://www.mydomain.com/dropbox_auth_start")
	except DropboxOAuth2Flow.CsrfException as e:
		raise e
	except DropboxOAuth2Flow.NotApprovedException as e:
		raise e
	except DropboxOAuth2Flow.ProviderException as e:
		raise e
	except Exception as e:
		raise e
	return HttpResponseRedirect('kineticjstest.html')

def testjsp(request):
    return render_to_response('testjsp.jsp')

def ajaxresponse(request):
	data={}
	data['producer']='samsung'
	data['model']='razr'
	data['price']=5
	return HttpResponse(json.dumps(data), content_type='application/json')

def simpletest(request):
	return render(request, 'simpletest.html')

def get_dropbox_auth_flow(session):
	app_key='p7b7j5v29v38bnk'
	app_secret='q83heifozbsktdy'
	return dropbox.client.DropboxOAuth2Flow(app_key,app_secret,"https://bioscheduler.com/polls/registerdropbox",session,"dropbox-auth-csrf-token")
#	return dropbox.client.DropboxOAuth2Flow(app_key,app_secret,"http://localhost:8000/polls/registerdropbox",session,"dropbox-auth-csrf-token")

def augmentlist(request):
	newsequences=request.POST['oligos']
	sequence_list=json.loads(newsequences)
	try:
		client=dropbox.client.DropboxClient(request.session['access_token'])
		folder_metadata=client.metadata('/')
		for obj in folder_metadata['contents']:
			if 'mime_type' in obj and obj['mime_type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
				absolute_path=obj['path']
				f=client.get_file(absolute_path)
				s=io.BytesIO(f.read())
				book=load_workbook(filename=s)
				for sheet in book:
					row=2
					index=0
					while sheet.cell(row=row,column=1).value!=None:
						index=sheet.cell(row=row,column=1).value
						row+=1
					for seq in sequence_list:
						index+=1
						sheet.cell(row=row,column=1).value=index
						sheet.cell(row=row,column=2).value=seq['name']
						sheet.cell(row=row,column=3).value=seq['nonComplementary']+seq['complementary']
						sheet.cell(row=row,column=5).value=seq['nonComplementary']+seq['complementary']
						sheet.cell(row=row,column=6).value=seq['nonComplementary']
						sheet.cell(row=row,column=7).value=seq['complementary']
						row+=1
				client.put_file(absolute_path,save_virtual_workbook(book),overwrite=True)
	except Exception as e:
		print(str(e))
		traceback.print_exc()
		raise e
	return HttpResponse("success", content_type='application/text')

#def retrieve_modules(request):
#	try:
#		client=dropbox.client.DropboxClient(request.session['access_token'])
#		module_metadata=client.metadata('/modules')
#		for obj in module_metadata['contents']:
#
#			data.append([obj['path'],1,'','','','loadinventory'])


def getlist(request):
	data=[]
	try:
		client=dropbox.client.DropboxClient(request.session['access_token'])
		folder_metadata=client.metadata('/')
		for obj in folder_metadata['contents']:
			#print(json.dumps(obj))
			if 'mime_type' in obj and obj['mime_type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
				absolute_path=obj['path']
				f=client.get_file(absolute_path)
				s=f.read()
				book=open_workbook(file_contents=s)
				for name in book.sheet_names():
					sheet=book.sheet_by_name(name)
					rows=sheet.nrows
					current_row=2
					while current_row<rows:
						row=sheet.row(current_row)
						name_column=1
						text_entry=1
						if sheet.cell_type(current_row,name_column)!=text_entry:
							break
						name=sheet.cell_value(current_row,name_column)
						sequence_column=2
						sequence=sheet.cell_value(current_row,sequence_column).replace(" ","")
						extension=""
						complementary=""
						try:
							extension_column=5
							extension=sheet.cell_value(current_row,extension_column)
							complementary_column=6
							complementary=sheet.cell_value(current_row,complementary_column)
						except:
							pass
						data.append([name,len(sequence),sequence.upper(),complementary.upper(),extension.upper(),''])
						current_row+=1
			else:
				data.append([obj['path'],obj['size'],'','','','kineticsequence'])
		try:
			inventory_metadata=client.metadata('/inventory')
			for obj in inventory_metadata['contents']:
				data.append([obj['path'],1,'','','','locations_module'])
			item_metadata=client.metadata('/items')
			for obj in item_metadata['contents']:
				data.append([obj['path'],1,'','','','item_display'])
		except:
			pass
	except Exception as e:
		print(str(e))
		traceback.print_exc()
		raise e
	return HttpResponse(json.dumps(data),content_type='application/json')


def kineticjstest(request):
	try:
		client=dropbox.client.DropboxClient(request.session['access_token'])
		folder_metadata=client.metadata('/')
	except:
		flow=get_dropbox_auth_flow(request.session)
		return HttpResponseRedirect(flow.start())
	return render(request, 'kineticjstest.html')
#	return render_to_response('kineticjstest.html', RequestContext(request))

def testjsplumb(request):
	return render(request, 'testjsplumb.html')

def saveitem(request):
	item_json=request.POST['item']
	item_name=request.POST['name']
	client=dropbox.client.DropboxClient(request.session['access_token'])
	client.put_file(item_name,io.StringIO(item_json),overwrite=True)
	return HttpResponse(item_json, content_type='application/json')

def savesequence(request):
	sequence_json=request.POST['sequence']
	sequence_name=request.POST['name']
	try:
		seq_handle=SeqIO.parse(io.StringIO(sequence_json),'json')
		for index, record in enumerate(seq_handle):
			newseq=SeqRecord(Seq(str(record.seq),IUPAC.unambiguous_dna),id=record.id)
			newseq.features=record.features
			output=io.StringIO()
			SeqIO.write(newseq, output, 'genbank')
	except Exception as e:
		print(str(e))
		traceback.print_exc()
		raise e
	client=dropbox.client.DropboxClient(request.session['access_token'])
	client.put_file(sequence_name,output,overwrite=True)
	return HttpResponse(sequence_json, content_type='application/json')

def deletesequence(request):
	sequence_path=request.POST['path']
	try:
		client=dropbox.client.DropboxClient(request.session['access_token'])
		client.file_delete(sequence_path)
	except Exception as e:
		print(str(e))
		traceback.print_exc()
		raise e
	return HttpResponse("success", content_type='application/html')

def getdropboxsequence(request,name):
	fileName,fileExtension=os.path.splitext(name)
	try:
		client=dropbox.client.DropboxClient(request.session['access_token'])
		with client.get_file(name) as f:
			s=f.read()
	except Exception as e:
		print(str(e),file=sys.stderr)
		raise e

	if fileExtension in ('.gbk','.gb'):
		try:
			seq=SeqIO.read(io.StringIO(s.decode("utf-8")), "genbank")
		except Exception as e:
			print(str(e),file=sys.stderr)
			raise e
		return seq
	elif fileExtension=='.seq':
		simple_seq=Seq("".join(s.decode("utf-8").split()))
		seq=SeqRecord(simple_seq)
		seq.id=fileName
		return seq
	else:
		print('impossible case')
	return seq


def align(request):
	names=request.POST.getlist('sequences[]')
	directions=request.POST.getlist('directions[]')
	fasta_string=""
	aligned_sequences={}
	try:
		for (name, direction) in zip(names, directions):
			seq=getdropboxsequence(request,name)
			fasta_string+=">"+name+"\n"
			print("direction"+str(direction))
			if direction=="1":
				print("forward")
				fasta_string+=str(seq.seq)+"\n"
			else:
				print("reverse")
				fasta_string+=reverse_complement(str(seq.seq))+"\n"
		cmd=Popen(['muscle'],stdout=PIPE,stdin=PIPE)
		stdout_data,stderr_data=cmd.communicate(input=fasta_string.encode("utf-8"))
		align=AlignIO.read(io.StringIO(stdout_data.decode('utf-8')),"fasta")
		for record in align:
			aligned_sequences[record.id]=str(record.seq)
	except Exception as e:
		print(str(e),file=sys.stderr)
		raise e
	return HttpResponse(json.dumps(aligned_sequences), content_type='application/json')

def getsequence(request):
	absolute_path=request.POST['name']
	fileName,fileExtension=os.path.splitext(absolute_path)
	seq=getdropboxsequence(request,absolute_path)
	if fileExtension in ('.gbk', '.gb'):
		output=io.StringIO()
		SeqIO.write(seq, output, "json")
		jsonstring=output.getvalue()
		outputstring=jsonstring[2:-2]
	elif fileExtension=='.seq':
		outputstring=json.dumps(seq)
	return HttpResponse(outputstring, content_type='application/json')

def getlocation(request):
	print('getlocation')
	absolute_path=request.POST['name']
	try:
		client=dropbox.client.DropboxClient(request.session['access_token'])
		with client.get_file(absolute_path) as f:
			s=f.read()
	except Exception as e:
		print(str(e),file=sys.stderr)
		raise e
	return HttpResponse(io.StringIO(s.decode("utf-8")), content_type='application/json')

def getitem(request):
	absolute_path=request.POST['name']
	print('getitem',file=sys.stderr)
	try:
		client=dropbox.client.DropboxClient(request.session['access_token'])
		with client.get_file(absolute_path) as f:
			s=f.read()
	except Exception as e:
		print(str(e),file=sys.stderr)
		raise e
	print(s.decode("utf-8"))
	return HttpResponse(io.StringIO(s.decode("utf-8")), content_type='application/json')
